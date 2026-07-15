#!/usr/bin/env python3
"""Programmatic quality control for the recalculated final Excel workbook."""

from __future__ import annotations

import argparse
import json
import math
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = PROJECT_ROOT / "results" / "notebook94_final_kalman_matlab_comparison"
DEFAULT_WORKBOOK = DEFAULT_OUT / "NathTrack_MATLAB_Python_Final_Comparison.xlsx"
EXPECTED_SHEETS = [
    "README", "Run_Metadata", "Framewise_Data", "Summary_Metrics", "Fixed_vs_Adaptive",
    "Bland_Altman", "Regression", "Alignment_Check", "Confidence_Analysis",
    "Outlier_Frames", "Figure_Data", "Data_Dictionary", "Manuscript_Summary", "Plots",
]
CSV_SHEETS = {
    "Framewise_Data": "Framewise_Data.csv",
    "Summary_Metrics": "Summary_Metrics.csv",
    "Fixed_vs_Adaptive": "Fixed_vs_Adaptive.csv",
    "Bland_Altman": "Bland_Altman.csv",
    "Regression": "Regression.csv",
    "Confidence_Analysis": "Confidence_Analysis.csv",
    "Outlier_Frames": "Outlier_Frames.csv",
    "Figure_Data": "Figure_Data.csv",
    "Data_Dictionary": "Data_Dictionary.csv",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--report", type=Path, default=DEFAULT_OUT / "workbook_QC_report.json")
    return parser.parse_args()


def equal_numeric(a: Any, b: Any, tolerance: float = 1e-10) -> bool:
    if a is None and (b is None or pd.isna(b)):
        return True
    try:
        return math.isclose(float(a), float(b), rel_tol=0.0, abs_tol=tolerance)
    except (TypeError, ValueError):
        return a == b


def main() -> None:
    args = parse_args()
    failures: list[str] = []
    checks: dict[str, Any] = {}
    if not args.workbook.exists():
        raise FileNotFoundError(args.workbook)

    with zipfile.ZipFile(args.workbook) as archive:
        bad = archive.testzip()
        checks["xlsx_zip_integrity"] = bad is None
        if bad is not None:
            failures.append(f"Corrupt XLSX member: {bad}")

    formula_wb = load_workbook(args.workbook, data_only=False, read_only=False)
    value_wb = load_workbook(args.workbook, data_only=True, read_only=False)
    checks["sheet_names"] = formula_wb.sheetnames
    if formula_wb.sheetnames != EXPECTED_SHEETS:
        failures.append(f"Expected sheets {EXPECTED_SHEETS}, got {formula_wb.sheetnames}")
    hidden = [ws.title for ws in formula_wb.worksheets if ws.sheet_state != "visible"]
    checks["hidden_sheets"] = hidden
    if hidden:
        failures.append(f"Hidden sheets found: {hidden}")
    embedded_images = len(formula_wb["Plots"]._images)
    checks["embedded_plot_images"] = embedded_images
    if embedded_images != 10:
        failures.append(f"Expected 10 embedded plot images, found {embedded_images}")

    sources = {sheet: pd.read_csv(args.input_dir / filename) for sheet, filename in CSV_SHEETS.items()}
    for sheet, source in sources.items():
        ws = formula_wb[sheet]
        workbook_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        if workbook_headers != list(source.columns):
            failures.append(f"{sheet}: headers/column count do not match source CSV")
        if ws.max_row - 1 != len(source):
            failures.append(f"{sheet}: workbook rows {ws.max_row-1} != source rows {len(source)}")
        zero_width = [col for col in range(1, ws.max_column + 1) if ws.column_dimensions[ws.cell(1, col).column_letter].width in (None, 0)]
        if zero_width:
            failures.append(f"{sheet}: columns without usable widths: {zero_width}")
    checks["source_row_counts"] = {sheet: len(source) for sheet, source in sources.items()}

    formula_errors: list[str] = []
    cached_errors: list[str] = []
    formula_count = 0
    literal_nan_count = 0
    for ws in formula_wb.worksheets:
        value_ws = value_wb[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if "#REF!" in value or "#NAME?" in value:
                        formula_errors.append(f"{ws.title}!{cell.coordinate}: {value}")
                    cached = value_ws[cell.coordinate].value
                    if isinstance(cached, str) and cached.startswith("#"):
                        cached_errors.append(f"{ws.title}!{cell.coordinate}: {cached}")
                if isinstance(value, str) and value.strip().lower() in {"nan", "na", "n/a"}:
                    literal_nan_count += 1
    checks["formula_count"] = formula_count
    checks["formula_reference_errors"] = formula_errors
    checks["cached_formula_errors"] = cached_errors
    checks["literal_nan_cells"] = literal_nan_count
    if formula_count == 0:
        failures.append("No visible formulas found")
    if formula_errors or cached_errors:
        failures.append("Formula errors detected")
    if literal_nan_count:
        failures.append(f"Found {literal_nan_count} literal NaN/NA cells; missing values must be blank")

    frame_source = sources["Framewise_Data"]
    frame_formula_ws = formula_wb["Framewise_Data"]
    frame_value_ws = value_wb["Framewise_Data"]
    header = {cell.value: cell.column for cell in frame_formula_ws[1]}
    source_spotchecks: list[dict[str, Any]] = []
    for source_idx in sorted({0, len(frame_source) // 2, len(frame_source) - 2, len(frame_source) - 1}):
        excel_row = source_idx + 2
        for column in [
            "frame_index_video", "MATLAB_ANG_deg", "Python_fixed_ANG_deg", "Python_adaptive_ANG_deg",
            "MATLAB_PEN_deg", "MATLAB_FL_mm", "adaptive_confidence_score", "adaptive_R_scale_angle",
            "hough_localmax_fallback_flag", "included_in_analysis",
        ]:
            excel_value = frame_value_ws.cell(excel_row, header[column]).value
            csv_value = frame_source.iloc[source_idx][column]
            ok = equal_numeric(excel_value, csv_value)
            source_spotchecks.append({"source_row": source_idx, "column": column, "excel": excel_value, "csv": None if pd.isna(csv_value) else csv_value, "match": ok})
            if not ok:
                failures.append(f"Framewise spotcheck mismatch row {source_idx}, column {column}")
    checks["csv_spotchecks"] = source_spotchecks

    metadata = json.loads((args.input_dir / "Run_Metadata.json").read_text(encoding="utf-8"))
    matlab_checks: list[dict[str, Any]] = []
    for item in metadata["matlab_source_spotchecks"]:
        source_idx = int(item["matlab_sample_index_zero_based"])
        excel_row = source_idx + 2
        for column in ["MATLAB_ANG_deg", "MATLAB_PEN_deg", "MATLAB_FL_mm"]:
            excel_value = frame_value_ws.cell(excel_row, header[column]).value
            mat_value = item[column]
            ok = equal_numeric(excel_value, mat_value)
            matlab_checks.append({"matlab_sample": source_idx, "column": column, "excel": excel_value, "mat_file_value": mat_value, "match": ok})
            if not ok:
                failures.append(f"MATLAB source mismatch sample {source_idx}, column {column}")
    checks["mat_file_spotchecks"] = matlab_checks

    formula_columns = [
        "fixed_minus_matlab_ANG_deg", "adaptive_minus_matlab_ANG_deg",
        "fixed_minus_matlab_PEN_deg", "adaptive_minus_matlab_PEN_deg",
        "fixed_minus_matlab_FL_mm", "adaptive_minus_matlab_FL_mm",
        "absolute_error_fixed_ANG_deg", "absolute_error_adaptive_ANG_deg",
        "absolute_error_fixed_PEN_deg", "absolute_error_adaptive_PEN_deg",
        "absolute_error_fixed_FL_mm", "absolute_error_adaptive_FL_mm",
    ]
    formula_cache_checks = []
    for row in [2, 2 + len(frame_source) // 2, len(frame_source)]:
        for column in formula_columns:
            formula = frame_formula_ws.cell(row, header[column]).value
            cached = frame_value_ws.cell(row, header[column]).value
            ok = isinstance(formula, str) and formula.startswith("=") and isinstance(cached, (int, float))
            formula_cache_checks.append({"cell": frame_formula_ws.cell(row, header[column]).coordinate, "formula": formula, "cached_value": cached, "valid": ok})
            if not ok:
                failures.append(f"Formula/cached value missing at Framewise_Data!{frame_formula_ws.cell(row, header[column]).coordinate}")
    checks["formula_cache_spotchecks"] = formula_cache_checks

    included_values = [frame_value_ws.cell(row, header["included_in_analysis"]).value for row in range(2, frame_value_ws.max_row + 1)]
    included_count = sum(value is True for value in included_values)
    checks["framewise_rows"] = frame_value_ws.max_row - 1
    checks["included_frames"] = included_count
    if frame_value_ws.max_row - 1 != metadata["video_frames"]:
        failures.append("Excel frame count does not match video frame count")
    if included_count != metadata["final_paired_frames"]:
        failures.append("Excel included-frame count does not match final paired-frame count")

    summary_ws = value_wb["Summary_Metrics"]
    summary_header = {cell.value: cell.column for cell in summary_ws[1]}
    units = {(summary_ws.cell(row, summary_header["variable"]).value, summary_ws.cell(row, summary_header["unit"]).value) for row in range(2, summary_ws.max_row + 1)}
    checks["summary_units"] = sorted([list(value) for value in units])
    if units != {("ANG", "deg"), ("PEN", "deg"), ("FL", "mm")}:
        failures.append(f"Unexpected Summary_Metrics units: {units}")

    numeric_type_failures = []
    for column in ["MATLAB_ANG_deg", "Python_fixed_ANG_deg", "MATLAB_FL_mm", "adaptive_confidence_score"]:
        for row in [2, 100, 1000]:
            value = frame_value_ws.cell(row, header[column]).value
            if not isinstance(value, (int, float)):
                numeric_type_failures.append(f"{column} row {row}: {type(value).__name__}")
    checks["numeric_type_failures"] = numeric_type_failures
    if numeric_type_failures:
        failures.append("Numeric cells stored as non-numeric values")

    formula_wb.close()
    value_wb.close()
    checks["failures"] = failures
    checks["passed"] = not failures
    args.report.write_text(json.dumps(checks, indent=2, default=str) + "\n", encoding="utf-8")
    print(json.dumps({"passed": not failures, "failures": failures, "report": str(args.report.resolve())}, indent=2))
    if failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
