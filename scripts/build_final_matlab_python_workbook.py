#!/usr/bin/env python3
"""Create the final co-author Excel workbook from validated flat artifacts.

Run this script with the bundled workspace Python that provides openpyxl.
The user explicitly requested openpyxl or xlsxwriter for workbook generation.
"""

from __future__ import annotations

import argparse
import importlib.metadata
import json
import math
import sys
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = PROJECT_ROOT / "results" / "notebook94_final_kalman_matlab_comparison"
DEFAULT_WORKBOOK = DEFAULT_OUT / "NathTrack_MATLAB_Python_Final_Comparison.xlsx"

EXPECTED_SHEETS = [
    "README",
    "Run_Metadata",
    "Framewise_Data",
    "Summary_Metrics",
    "Fixed_vs_Adaptive",
    "Bland_Altman",
    "Regression",
    "Alignment_Check",
    "Confidence_Analysis",
    "Outlier_Frames",
    "Figure_Data",
    "Data_Dictionary",
    "Manuscript_Summary",
    "Plots",
]

NAVY = "1F4E78"
BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
PALE_GREEN = "E2F0D9"
PALE_AMBER = "FFF2CC"
PALE_RED = "F4CCCC"
WHITE = "FFFFFF"
GREY = "666666"
THIN_GREY = Side(style="thin", color="D9E1F2")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_WORKBOOK)
    return parser.parse_args()


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def title_band(ws, title: str, subtitle: str | None = None, width: int = 8) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        ws.cell(2, 1, subtitle)
        ws.cell(2, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        ws.cell(2, 1).font = Font(name="Aptos", size=10, italic=True, color=GREY)
        ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 34
        return 4
    return 3


def style_header(ws, row: int, start_col: int, end_col: int) -> None:
    for cell in ws.iter_cols(min_col=start_col, max_col=end_col, min_row=row, max_row=row):
        c = cell[0]
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[row].height = 42


def add_table(ws, start_row: int, end_row: int, end_col: int, name: str) -> None:
    if end_row <= start_row or end_col <= 0:
        return
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def number_format_for_header(header: str) -> str:
    lower = header.lower()
    if lower.endswith("_n") or "count" in lower or lower.startswith("frame_index") or lower.endswith("_frames"):
        return "0"
    if "percent" in lower:
        return "0.00"
    if "time_seconds" in lower:
        return "0.000000"
    if any(token in lower for token in ["correlation", "pearson", "spearman", "r_squared", "confidence", "fraction"]):
        return "0.0000"
    if any(token in lower for token in ["angle", "ang", "pen", "deg", "slope", "intercept", "bias", "error", "mae", "rmse", "loa", "covariance", "gain", "scale", "value", "mean", "median", "width", "threshold", "jump", "fl_mm", "fl_px"]):
        return "0.0000"
    return "General"


def set_column_widths(ws, headers: list[str], rows: Iterable[Iterable[Any]], sample_limit: int = 200) -> None:
    samples = list(rows)[:sample_limit]
    for col_idx, header in enumerate(headers, 1):
        values = [str(header)]
        for row in samples:
            value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if value is not None:
                values.append(str(value))
        max_len = max(len(value) for value in values)
        if any(token in header.lower() for token in ["reason", "definition", "caution", "command", "path", "source", "interpretation"]):
            width = min(max(max_len + 2, 22), 55)
        else:
            width = min(max(max_len + 2, 10), 22)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def dataframe_sheet(
    wb: Workbook,
    name: str,
    df: pd.DataFrame,
    *,
    table_name: str,
    freeze: str = "A2",
) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    headers = [str(column) for column in df.columns]
    rows = [[clean_value(value) for value in row] for row in df.itertuples(index=False, name=None)]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, 1, 1, len(headers))
    add_table(ws, 1, len(rows) + 1, len(headers), table_name)
    ws.freeze_panes = freeze
    set_column_widths(ws, headers, rows)
    for col_idx, header in enumerate(headers, 1):
        number_format = number_format_for_header(header)
        if number_format != "General":
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(rows) + 1):
                cell[0].number_format = number_format
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"


def add_framewise_formulas(ws) -> None:
    header_map = {cell.value: cell.column for cell in ws[1]}
    formulas = [
        ("fixed_minus_matlab_ANG_deg", "Python_fixed_ANG_deg", "MATLAB_ANG_deg", "difference"),
        ("adaptive_minus_matlab_ANG_deg", "Python_adaptive_ANG_deg", "MATLAB_ANG_deg", "difference"),
        ("fixed_minus_matlab_PEN_deg", "Python_fixed_PEN_deg", "MATLAB_PEN_deg", "difference"),
        ("adaptive_minus_matlab_PEN_deg", "Python_adaptive_PEN_deg", "MATLAB_PEN_deg", "difference"),
        ("fixed_minus_matlab_FL_mm", "Python_fixed_FL_mm", "MATLAB_FL_mm", "difference"),
        ("adaptive_minus_matlab_FL_mm", "Python_adaptive_FL_mm", "MATLAB_FL_mm", "difference"),
        ("absolute_error_fixed_ANG_deg", "fixed_minus_matlab_ANG_deg", None, "absolute"),
        ("absolute_error_adaptive_ANG_deg", "adaptive_minus_matlab_ANG_deg", None, "absolute"),
        ("absolute_error_fixed_PEN_deg", "fixed_minus_matlab_PEN_deg", None, "absolute"),
        ("absolute_error_adaptive_PEN_deg", "adaptive_minus_matlab_PEN_deg", None, "absolute"),
        ("absolute_error_fixed_FL_mm", "fixed_minus_matlab_FL_mm", None, "absolute"),
        ("absolute_error_adaptive_FL_mm", "adaptive_minus_matlab_FL_mm", None, "absolute"),
    ]
    for target, source_a, source_b, kind in formulas:
        target_col = header_map[target]
        a = get_column_letter(header_map[source_a])
        b = get_column_letter(header_map[source_b]) if source_b else None
        for row in range(2, ws.max_row + 1):
            if kind == "difference":
                formula = f'=IF(OR({a}{row}="",{b}{row}=""),"",{a}{row}-{b}{row})'
            else:
                formula = f'=IF({a}{row}="","",ABS({a}{row}))'
            ws.cell(row, target_col, formula)
            ws.cell(row, target_col).number_format = "0.0000"
        ws.cell(1, target_col).comment = Comment(
            "Visible Excel formula; the same value was independently calculated in Python and checked before workbook creation.",
            "Codex scientific validation",
        )

    included_col = get_column_letter(header_map["included_in_analysis"])
    fallback_col = get_column_letter(header_map["hough_localmax_fallback_flag"])
    conf_col = get_column_letter(header_map["adaptive_confidence_score"])
    ws.conditional_formatting.add(
        f"{included_col}2:{included_col}{ws.max_row}",
        FormulaRule(formula=[f"NOT({included_col}2)"], fill=PatternFill("solid", fgColor=PALE_RED)),
    )
    ws.conditional_formatting.add(
        f"{fallback_col}2:{fallback_col}{ws.max_row}",
        FormulaRule(formula=[f"{fallback_col}2=TRUE"], fill=PatternFill("solid", fgColor=PALE_AMBER)),
    )
    ws.conditional_formatting.add(
        f"{conf_col}2:{conf_col}{ws.max_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=0.5, mid_color="FFEB84", end_type="num", end_value=1, end_color="63BE7B"),
    )


def add_bland_altman_formulas(ws) -> None:
    header_map = {cell.value: cell.column for cell in ws[1]}
    mat_col = get_column_letter(header_map["MATLAB_value"])
    py_col = get_column_letter(header_map["Python_value"])
    mean_col = header_map["bland_altman_mean"]
    diff_col = header_map["difference_python_minus_matlab"]
    for row in range(2, ws.max_row + 1):
        ws.cell(row, mean_col, f'=IF(OR({mat_col}{row}="",{py_col}{row}=""),"",AVERAGE({mat_col}{row},{py_col}{row}))')
        ws.cell(row, diff_col, f'=IF(OR({mat_col}{row}="",{py_col}{row}=""),"",{py_col}{row}-{mat_col}{row})')
        ws.cell(row, mean_col).number_format = "0.0000"
        ws.cell(row, diff_col).number_format = "0.0000"
    for column in (mean_col, diff_col):
        ws.cell(1, column).comment = Comment(
            "Visible frame-level Excel formula, independently checked against Python analysis output.",
            "Codex scientific validation",
        )


def flatten_json(value: Any, prefix: str = "") -> list[tuple[str, str, Any]]:
    rows: list[tuple[str, str, Any]] = []
    if isinstance(value, dict):
        for key, item in value.items():
            section = prefix.split(".", 1)[0] if prefix else str(key)
            path = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(item, (dict, list)):
                rows.extend(flatten_json(item, path))
            else:
                rows.append((section, path, item))
    elif isinstance(value, list):
        for idx, item in enumerate(value):
            path = f"{prefix}[{idx}]"
            if isinstance(item, (dict, list)):
                rows.extend(flatten_json(item, path))
            else:
                rows.append((prefix.split(".", 1)[0], path, item))
    else:
        rows.append((prefix.split(".", 1)[0], prefix, value))
    return rows


def build_readme(wb: Workbook, metadata: dict[str, Any]) -> None:
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    row = title_band(
        ws,
        "NathTrack–MATLAB final comparison",
        "Publication-oriented technical validation workbook; all frame-level measurements and diagnostics remain visible.",
        6,
    )
    items = [
        ("Purpose", metadata["purpose"]),
        ("Analysis date (UTC)", metadata["analysis_timestamp_utc"]),
        ("Dataset / video", metadata["video_path"]),
        ("MATLAB input", metadata["matlab_result_path"]),
        ("Python input", metadata["python_npz_path"]),
        ("ROI input", metadata["roi_path"]),
        ("Code version", f"Branch {metadata['git_branch']}; commit {metadata['git_commit']}"),
        ("Final temporal alignment", f"Python physical-frame offset {metadata['final_python_offset_frames']}. {metadata['alignment_basis']}"),
        ("Final paired frames", metadata["final_paired_frames"]),
        (
            "Python seed initialization",
            f"{metadata['seed_initialization']['seed_frames']} early frames; seed-only range "
            f"{metadata['seed_initialization']['seed_only_angle_min_deg']:g}–{metadata['seed_initialization']['seed_only_angle_max_deg']:g}°, "
            f"selected {metadata['seed_initialization']['selected_seed_alpha_deg']:g}° "
            f"({metadata['seed_initialization']['selected_cluster_id']}). Per-frame Hough remained "
            f"{metadata['seed_initialization']['per_frame_hough_angle_min_deg']:g}–{metadata['seed_initialization']['per_frame_hough_angle_max_deg']:g}°. "
            f"{metadata['seed_initialization']['rationale']}"
        ),
        ("ANG definition", "Fascicle orientation in image coordinates using atan2(-Δy, Δx), normalized to [-90°, 90°). Positive values describe the tracked downward-left fascicle after endpoint normalization."),
        ("PEN definition", "Pennation angle = fascicle ANG minus deep-aponeurosis angle, in degrees."),
        ("FL definition", "Euclidean distance between intersections of the final fascicle line with the superficial and deep aponeuroses. Reported in mm using the validated scale."),
        ("Superficial/deep coordinates", "One-based MATLAB-compatible image coordinates: x increases right and y increases downward. Aponeurosis endpoints are left-to-right; fascicle endpoints are labelled superficial/deep explicitly."),
        ("Fixed-R", "Python MATLAB-like two-state Kalman filter with constant measurement covariance."),
        ("Adaptive anisotropic-R", "The same filter and measurements, with per-frame angle and length-side covariance scales derived separately from confidence diagnostics."),
        ("Important caution", "MATLAB is a comparison reference, not an error-free ground truth. No data were interpolated, extrapolated, temporally shifted, or smoothed for plotting beyond the implemented Kalman/RTS pipeline itself."),
        ("Equivalence", metadata["equivalence_testing"]),
        ("Generalisability", "One video sequence cannot establish population-level performance or statistical equivalence. Results are a within-sequence technical comparison."),
        ("Missing values", "Blank cells represent unavailable values. The final Python frame has no MATLAB counterpart and is retained with included_in_analysis=FALSE and an explicit exclusion reason."),
    ]
    ws.cell(row, 1, "Topic"); ws.cell(row, 2, "Detail")
    style_header(ws, row, 1, 2)
    for topic, detail in items:
        row += 1
        ws.cell(row, 1, topic)
        ws.cell(row, 2, clean_value(detail))
        ws.cell(row, 1).font = Font(name="Aptos", size=10, bold=True, color=NAVY)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(22, min(68, 15 * (str(detail).count("\n") + 2)))
    ws.column_dimensions["A"].width = 29
    ws.column_dimensions["B"].width = 115
    ws.freeze_panes = "A5"
    add_table(ws, 4, row, 2, "README_Table")


def build_metadata(wb: Workbook, metadata: dict[str, Any]) -> None:
    ws = wb.create_sheet("Run_Metadata")
    ws.sheet_view.showGridLines = False
    start = title_band(ws, "Run metadata", "Paths, commands, code state, package versions, counts, units, and algorithm parameters.", 5)
    headers = ["section", "metadata_key", "value"]
    for col, header in enumerate(headers, 1):
        ws.cell(start, col, header)
    style_header(ws, start, 1, 3)
    rows = flatten_json(metadata)
    for row_idx, values in enumerate(rows, start + 1):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, clean_value(value))
        ws.cell(row_idx, 3).alignment = Alignment(wrap_text=True, vertical="top")
    add_table(ws, start, start + len(rows), 3, "RunMetadataTable")
    ws.freeze_panes = f"A{start+1}"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 105


def build_alignment(wb: Workbook, alignment: pd.DataFrame, physical: pd.DataFrame, metadata: dict[str, Any]) -> None:
    ws = wb.create_sheet("Alignment_Check")
    ws.sheet_view.showGridLines = False
    start = title_band(ws, "Temporal alignment check", metadata["alignment_basis"], 12)
    ws.cell(start, 1, "Final-output sensitivity (offsets −3 to +3)")
    ws.cell(start, 1).font = Font(name="Aptos", size=11, bold=True, color=NAVY)
    header_row = start + 1
    headers = list(alignment.columns)
    for col, header in enumerate(headers, 1): ws.cell(header_row, col, header)
    style_header(ws, header_row, 1, len(headers))
    for r_idx, row in enumerate(alignment.itertuples(index=False, name=None), header_row + 1):
        for c_idx, value in enumerate(row, 1): ws.cell(r_idx, c_idx, clean_value(value))
    end = header_row + len(alignment)
    add_table(ws, header_row, end, len(headers), "AlignmentSensitivityTable")

    evidence_title = end + 3
    ws.cell(evidence_title, 1, "Independent physical-frame evidence")
    ws.cell(evidence_title, 1).font = Font(name="Aptos", size=11, bold=True, color=NAVY)
    evidence_header = evidence_title + 1
    p_headers = list(physical.columns)
    for col, header in enumerate(p_headers, 1): ws.cell(evidence_header, col, header)
    style_header(ws, evidence_header, 1, len(p_headers))
    for r_idx, row in enumerate(physical.itertuples(index=False, name=None), evidence_header + 1):
        for c_idx, value in enumerate(row, 1): ws.cell(r_idx, c_idx, clean_value(value))
    add_table(ws, evidence_header, evidence_header + len(physical), len(p_headers), "AlignmentPhysicalEvidenceTable")
    ws.freeze_panes = f"A{header_row+1}"
    set_column_widths(ws, headers, alignment.itertuples(index=False, name=None))


def build_manuscript(wb: Workbook, manuscript: dict[str, Any]) -> None:
    ws = wb.create_sheet("Manuscript_Summary")
    ws.sheet_view.showGridLines = False
    row = title_band(ws, "Manuscript summary", "Concise co-author wording; factual observations are separated from interpretation and caveats.", 8)
    sections = [
        ("Suggested Methods wording", manuscript["Methods_suggested"], PALE_BLUE),
        ("Suggested Results wording", manuscript["Results_suggested"], PALE_GREEN),
        ("Interpretation", manuscript["Interpretation"], PALE_AMBER),
        ("Explicit caveats", "\n".join(f"• {item}" for item in manuscript["Caveats"]), PALE_RED),
    ]
    for heading, body, color in sections:
        ws.cell(row, 1, heading)
        ws.cell(row, 1).font = Font(name="Aptos", size=11, bold=True, color=NAVY)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=color)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        ws.cell(row, 1, body)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        estimated_lines = body.count("\n") + math.ceil(len(body) / 150)
        ws.row_dimensions[row].height = min(210, max(55, 15 * (estimated_lines + 1)))
        row += 2
    ws.column_dimensions["A"].width = 24
    for col in range(2, 9): ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A4"


def build_plots(wb: Workbook, plots_dir: Path) -> None:
    ws = wb.create_sheet("Plots")
    ws.sheet_view.showGridLines = False
    title_band(ws, "Main validation plots", "Embedded PNG copies; figure-ready PNG and SVG files are retained separately with numerical source data in Figure_Data.", 16)
    plot_names = [
        "timeseries_ANG.png", "timeseries_PEN.png", "timeseries_FL.png",
        "bland_altman_grid.png", "scatter_identity_grid.png",
        "absolute_error_over_time_grid.png", "confidence_r_scale_over_time.png",
        "error_vs_confidence_grid.png", "fixed_vs_adaptive_absolute_error_grid.png",
        "temporal_offset_sensitivity_grid.png",
    ]
    anchors = ["A4", "A27", "A50", "A73", "A128", "A183", "A225", "A267", "A309", "A345"]
    for name, anchor in zip(plot_names, anchors):
        path = plots_dir / name
        if not path.exists():
            continue
        image = XLImage(path)
        if image.width > 1120:
            ratio = 1120 / image.width
            image.width = int(image.width * ratio)
            image.height = int(image.height * ratio)
        ws.add_image(image, anchor)
    for col in range(1, 17): ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = "A4"


def verify_created_workbook(path: Path, source_counts: dict[str, int]) -> None:
    wb = load_workbook(path, data_only=False, read_only=False)
    if wb.sheetnames != EXPECTED_SHEETS:
        raise AssertionError(f"Unexpected sheets: {wb.sheetnames}")
    frame_ws = wb["Framewise_Data"]
    if frame_ws.max_row - 1 != source_counts["Framewise_Data"]:
        raise AssertionError("Framewise_Data row count differs from source CSV.")
    header = {cell.value: cell.column for cell in frame_ws[1]}
    for name in ["fixed_minus_matlab_ANG_deg", "absolute_error_adaptive_FL_mm"]:
        value = frame_ws.cell(2, header[name]).value
        if not isinstance(value, str) or not value.startswith("="):
            raise AssertionError(f"Expected visible Excel formula in {name}.")
    for sheet, expected in source_counts.items():
        if sheet in {"Alignment_Check"}:
            continue
        if wb[sheet].max_row - 1 != expected:
            raise AssertionError(f"{sheet} row count differs from source CSV.")
    wb.close()


def main() -> None:
    args = parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    metadata = json.loads((args.input_dir / "Run_Metadata.json").read_text(encoding="utf-8"))
    metadata["workbook_builder_environment"] = {
        "python_executable": sys.executable,
        "python_version": sys.version,
        "openpyxl_version": importlib.metadata.version("openpyxl"),
        "pandas_version": pd.__version__,
    }
    manuscript = json.loads((args.input_dir / "Manuscript_Summary.json").read_text(encoding="utf-8"))
    csv_map = {
        "Framewise_Data": "Framewise_Data.csv",
        "Summary_Metrics": "Summary_Metrics.csv",
        "Fixed_vs_Adaptive": "Fixed_vs_Adaptive.csv",
        "Bland_Altman": "Bland_Altman.csv",
        "Regression": "Regression.csv",
        "Confidence_Analysis": "Confidence_Analysis.csv",
        "Outlier_Frames": "Outlier_Frames.csv",
        "Figure_Data": "Figure_Data.csv",
        "Data_Dictionary": "Data_Dictionary.csv",
    }
    frames: dict[str, pd.DataFrame] = {name: pd.read_csv(args.input_dir / filename) for name, filename in csv_map.items()}
    alignment = pd.read_csv(args.input_dir / "Alignment_Check.csv")
    physical = pd.read_csv(args.input_dir / "Alignment_Physical_Evidence.csv")

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    build_readme(wb, metadata)
    build_metadata(wb, metadata)
    for sheet_name in ["Framewise_Data", "Summary_Metrics", "Fixed_vs_Adaptive", "Bland_Altman", "Regression"]:
        dataframe_sheet(wb, sheet_name, frames[sheet_name], table_name=f"{sheet_name.replace('_', '')}Table", freeze="E2" if sheet_name == "Framewise_Data" else "A2")
    add_framewise_formulas(wb["Framewise_Data"])
    add_bland_altman_formulas(wb["Bland_Altman"])
    build_alignment(wb, alignment, physical, metadata)
    for sheet_name in ["Confidence_Analysis", "Outlier_Frames", "Figure_Data", "Data_Dictionary"]:
        dataframe_sheet(wb, sheet_name, frames[sheet_name], table_name=f"{sheet_name.replace('_', '')}Table")
    build_manuscript(wb, manuscript)
    build_plots(wb, args.input_dir / "plots")
    wb.save(args.output)
    verify_created_workbook(args.output, {name: len(df) for name, df in frames.items()})
    print(json.dumps({"workbook": str(args.output.resolve()), "sheets": wb.sheetnames, "source_rows": {name: len(df) for name, df in frames.items()}}, indent=2))


if __name__ == "__main__":
    main()
